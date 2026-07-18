import os # to check whether a cache file already exists
import time # to pause briefly between per-player API calls
import pandas as pd # for merging and organizing the dataframes
from nba_api.stats.endpoints import leaguedashplayerbiostats # league-wide bio stats (height, weight, country, age, team)
from nba_api.stats.endpoints import leaguedashplayerstats # league-wide game stats (points, shooting, rebounds, etc.)
from nba_api.stats.endpoints import commonplayerinfo # per-player info (position, birthdate, debut year)
from openpyxl import load_workbook # to re-open and format the saved xlsx
from openpyxl.styles import Font, Alignment # to control font (bold) and alignment (centered)

SEASON = "2025-26" # season to query, used everywhere below
SEASON_TYPE = "Regular Season" # season type to query, used everywhere below
CACHE_FILE = "nba_players_25_26_regular_season_wide_data_cache.csv" # caches only the SLOW, stable data (bio + position + birthdate)

# official NBA position acronyms, used to replace the long-form strings returned by commonplayerinfo
POSITION_ACRONYMS = {
    "Center": "C",
    "Center-Forward": "C-F",
    "Forward": "F",
    "Forward-Center": "F-C",
    "Forward-Guard": "F-G",
    "Guard": "G",
    "Guard-Forward": "G-F",
}

# --- Part A: bio/personal data (slow to fetch, but doesn't change during the season -> cached) ---
if os.path.exists(CACHE_FILE): # check if we already have this data saved locally
    print(f"Loading cached bio data from {CACHE_FILE} (no slow API calls needed)...")
    bio_merged = pd.read_csv(CACHE_FILE) # load the previously fetched data instantly
else:
    print("No bio cache found, fetching bio data from the NBA API (this will take a few minutes)...")

    # --- Step 1: bio data (height, weight, country, age, team) for every player in the season ---
    bio_stats = leaguedashplayerbiostats.LeagueDashPlayerBioStats( # call the bio stats endpoint
        season=SEASON, # season to query
        season_type_all_star=SEASON_TYPE, # regular season only
        timeout=60 # wait up to 60s for a response
    )
    bio_df = bio_stats.get_data_frames()[0] # extract the dataframe from the response
    bio_df = bio_df.drop(columns=[ # drop columns that also exist in the game-stats
        "GP", "PTS", "REB", "AST", # endpoint, to avoid duplicate column names on merge
        "NET_RATING", "OREB_PCT", "DREB_PCT", "USG_PCT", "TS_PCT", "AST_PCT", # unused advanced metrics
    ])

    # --- Step 2: loop through each player to get position, birthdate, and NBA debut year ---
    # (this data isn't available league-wide, so we query it player by player, with a short pause each time)
    extra_rows = [] # will store one dict per player
    total_players = len(bio_df) # for progress tracking
    for i, player_id in enumerate(bio_df["PLAYER_ID"], start=1): # loop through every player id
        try:
            info = commonplayerinfo.CommonPlayerInfo(player_id=player_id, timeout=60).get_data_frames()[0] # fetch player info
            extra_rows.append({ # save the fields we need
                "PLAYER_ID": player_id,
                "BIRTHDATE": info.loc[0, "BIRTHDATE"], # date of birth
                "POSITION": info.loc[0, "POSITION"], # playing position
                "FROM_YEAR": info.loc[0, "FROM_YEAR"], # first NBA season (debut year)
            })
        except Exception as e: # if a single player's request fails, skip it
            print(f"Skipped player_id {player_id}: {e}")
            extra_rows.append({"PLAYER_ID": player_id, "BIRTHDATE": None, "POSITION": None, "FROM_YEAR": None})
        if i % 50 == 0 or i == total_players: # print progress every 50 players
            print(f"Processed {i}/{total_players} players")
        time.sleep(0.6) # small pause to avoid overloading the API

    extra_df = pd.DataFrame(extra_rows) # convert the list of dicts into a dataframe
    bio_merged = bio_df.merge(extra_df, on="PLAYER_ID", how="left") # merge the per-player info into the bio table

    bio_merged.to_csv(CACHE_FILE, index=False) # save the bio data so we never re-fetch it
    print(f"Bio cache saved to {CACHE_FILE}")

# --- Part B: game stats (fast single call, but changes every game -> always fetched fresh) ---
# Only raw, non-derivable columns are requested here. GP, PTS, REB, and the *_PCT columns are
# intentionally left out because they can all be calculated afterwards from the columns below.
print("Fetching latest game stats from the NBA API...")
season_stats = leaguedashplayerstats.LeagueDashPlayerStats( # call the game stats endpoint
    season=SEASON, # season to query
    season_type_all_star=SEASON_TYPE, # regular season only
    per_mode_detailed="Totals", # season totals, not per-game averages
    timeout=60 # wait up to 60s for a response
)
stats_df = season_stats.get_data_frames()[0][[ # keep only the columns we need
    "PLAYER_ID", "W", "L", "MIN", # games played, wins, losses, minutes
    "FGM", "FGA", # field goals made/attempted (used to derive 2-point shots)
    "FG3M", "FG3A", # 3-pointers made/attempted
    "FTM", "FTA", # free throws made/attempted (= 1-point shots)
    "OREB", "DREB", # offensive/defensive rebounds
    "AST", "TOV", "STL", "BLK", "PF", # assists, turnovers, steals, blocks, fouls
]]

# --- Part C: merge bio (cached) + stats (fresh) into one table ---
merged = bio_merged.merge(stats_df, on="PLAYER_ID", how="left") # join on PLAYER_ID

# --- Part D: calculate derived columns (cheap, no API calls, safe to re-run anytime) ---
merged["SEASON"] = SEASON # add season as its own column
merged["SEASON_TYPE"] = SEASON_TYPE # add season type as its own column
merged["HEIGHT_M"] = (merged["PLAYER_HEIGHT_INCHES"] * 0.0254).round(3) # height in meters (for BMI calc in R later)
merged["WEIGHT_KG"] = (merged["PLAYER_WEIGHT"].astype(float) * 0.453592).round(1) # weight from lb to kg
current_season_start_year = int(SEASON[:4]) # e.g. 2025 from "2025-26"
merged["EXPERIENCE"] = current_season_start_year - merged["FROM_YEAR"].astype("Int64") + 1 # seasons since NBA debut
merged["BIRTHDATE"] = pd.to_datetime(merged["BIRTHDATE"]).dt.strftime("%Y-%m-%d") # keep only yyyy-mm-dd, drop the time part
merged["POSITION"] = merged["POSITION"].map(POSITION_ACRONYMS).fillna(merged["POSITION"]) # long-form -> official acronym

# 2-point shots aren't returned directly by the API, so derive them: field goals minus 3-pointers
merged["TWO_POINT_MADE"] = merged["FGM"] - merged["FG3M"]
merged["TWO_POINT_ATTEMPTED"] = merged["FGA"] - merged["FG3A"]

# --- Part E: select and order columns to mirror nba.com/stats/players/traditional layout ---
players_data = merged[[
    "SEASON", "SEASON_TYPE", "PLAYER_ID", "PLAYER_NAME", "TEAM_ABBREVIATION", "POSITION", # identity data
    "HEIGHT_M", "WEIGHT_KG", "COUNTRY", "AGE", "BIRTHDATE", "EXPERIENCE", # bio data
    "MIN", "W", "L", # playing time / team record
    "FTM", "FTA", # 1-point shots (free throws)
    "TWO_POINT_MADE", "TWO_POINT_ATTEMPTED", # 2-point shots (derived)
    "FG3M", "FG3A", # 3-point shots
    "OREB", "DREB", # rebounds
    "AST", "TOV", "STL", "BLK", "PF", # other stats
]].rename(columns={
    "TEAM_ABBREVIATION": "TEAM",
    "HEIGHT_M": "HEIGHT_M", "WEIGHT_KG": "WEIGHT_KG", "AGE": "AGE_COMPLETED_YEARS",
    "MIN": "MINUTES_PLAYED", "W": "WINS", "L": "LOSSES",
    "FTM": "1_POINT_MADE", "FTA": "1_POINT_ATTEMPTED",
    "TWO_POINT_MADE": "2_POINT_MADE", "TWO_POINT_ATTEMPTED": "2_POINT_ATTEMPTED",
    "FG3M": "3_POINT_MADE", "FG3A": "3_POINT_ATTEMPTED",
    "OREB": "OFFENSIVE_REBOUNDS", "DREB": "DEFENSIVE_REBOUNDS",
    "AST": "ASSISTS", "TOV": "TURNOVERS", "STL": "STEALS", "BLK": "BLOCKS", "PF": "PERSONAL_FOULS",
})

print(f"{len(players_data)} players found") # show how many players were retrieved
print(players_data) # preview the dataframe in the terminal

players_data.columns = players_data.columns.str.lower() # convert column headers to lowercase

players_data.to_csv("nba_players_25_26_regular_season_wide_data.csv", index=False) # save as CSV, no row index column

excel_path = "nba_players_25_26_regular_season_wide_data.xlsx" # define the Excel file path
players_data.to_excel(excel_path, index=False, sheet_name="Players") # save as Excel, no row index column

# --- Custom Excel formatting starts here ---
wb = load_workbook(excel_path) # reopen the saved Excel file
ws = wb["Players"] # select the "Players" sheet

plain_font = Font(bold=False) # define a non-bold font style (for data rows)
bold_font = Font(bold=True) # define a bold font style (for header row)
center_align = Alignment(horizontal="center", vertical="center") # define centered alignment (horizontal + vertical)

for row_index, row in enumerate(ws.iter_rows(), start=1): # loop through every row, tracking its row number (starts at 1)
    for cell in row: # loop through every cell in that row
        cell.font = bold_font if row_index == 1 else plain_font # bold only for row 1 (header), plain for the rest
        cell.alignment = center_align # apply centered alignment
        cell.border = None # explicitly remove any border (openpyxl default has none, but this ensures it)

ws.freeze_panes = "A2" # freeze the top row (everything above row 2 stays fixed when scrolling)

for column_cells in ws.columns: # loop through each column
    max_length = max(len(str(cell.value)) for cell in column_cells) # find the longest value in that column
    col_letter = column_cells[0].column_letter # get the column's letter (A, B, C...)
    ws.column_dimensions[col_letter].width = max_length + 2 # set width = longest text + small padding

wb.save(excel_path) # save the formatted Excel file, overwriting the previous version
print("Saved both CSV and formatted Excel files") # confirm completion