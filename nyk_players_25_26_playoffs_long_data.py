# https://www.nba.com/playoffs/2026

import time # to pause briefly between per-game API calls
import pandas as pd # for merging and organizing the dataframes
from nba_api.stats.endpoints import teamgamelog # team-level game log, used to work out round/game number
from nba_api.stats.endpoints import boxscoretraditionalv3 # full per-game box score, includes DNP/inactive players too
from openpyxl import load_workbook # to re-open and format the saved xlsx
from openpyxl.styles import Font, Alignment # to control font (bold) and alignment (centered)

SEASON = "2025-26" # season to query, used everywhere below
SEASON_TYPE = "Playoffs" # we only want playoff games
TEAM = "NYK" # New York Knicks (2025-26 NBA champions)
TEAM_ID = 1610612752 # official nba.com team id for the Knicks

# NBA playoff round names, in chronological order. The Knicks' opponent changes exactly once
# per round (First Round -> Conf. Semis -> East Finals -> NBA Finals), so we use each change
# of opponent in the team's game log to detect when a new round starts.
ROUND_NAMES = ["First Round", "Conference Semifinals", "Eastern Conference Finals", "NBA Finals"]

# --- Part A: team game log, only used to work out round_name / game_number / opponent per game ---
print("Fetching Knicks playoff team game log from the NBA API...")
team_games = teamgamelog.TeamGameLog( # call the team game log endpoint
    team_id=TEAM_ID, # New York Knicks
    season=SEASON, # season to query
    season_type_all_star=SEASON_TYPE, # playoffs only
    timeout=60 # wait up to 60s for a response
).get_data_frames()[0]

team_games["GAME_DATE"] = pd.to_datetime(team_games["GAME_DATE"], format="%b %d, %Y") # e.g. "APR 18, 2026" -> real date
team_games = team_games.sort_values("GAME_DATE").reset_index(drop=True) # now this really is chronological order
team_games["OPPONENT"] = team_games["MATCHUP"].str.split().str[-1] # last token of "NYK vs. ATL" / "NYK @ ATL" is the opponent

round_names_col = [] # will hold one round name per game, in order
round_index = 0 # 0 = not started yet, 1 = first round, etc.
previous_opponent = None # opponent faced in the previous game
for opponent in team_games["OPPONENT"]: # walk through the games in chronological order
    if opponent != previous_opponent: # a new opponent means a new round has started
        round_index += 1
        previous_opponent = opponent
    round_name = ROUND_NAMES[round_index - 1] if round_index - 1 < len(ROUND_NAMES) else f"Round {round_index}" # fallback just in case
    round_names_col.append(round_name)

team_games["ROUND_NAME"] = round_names_col # e.g. "First Round", "NBA Finals"...
team_games["GAME_NUMBER"] = range(1, len(team_games) + 1) # 1..19, correlative across the whole playoff run

print("Detected rounds (sanity check before trusting the labels below):")
print(team_games[["GAME_DATE", "OPPONENT", "ROUND_NAME", "GAME_NUMBER"]]) # quick manual check of the round detection

games_info = team_games[["Game_ID", "OPPONENT", "ROUND_NAME", "GAME_NUMBER"]].rename(columns={"Game_ID": "GAME_ID"})

# --- Part B: full Knicks box score (active AND inactive/DNP players) for every one of the 19 games ---
# NOTE: BoxScoreTraditionalV2 is deprecated for the 2025-26 season, so we use V3 instead.
# V3 uses camelCase field names (personId, teamId, minutes, comment...) instead of V2's UPPER_SNAKE_CASE.
def minutes_to_decimal(min_str):
    """Convert a box-score minutes value into decimal minutes. Handles both "MM:SS" and empty/None (DNP)."""
    if min_str is None or pd.isna(min_str) or str(min_str).strip() == "": # DNP players have no minutes at all
        return 0.0
    min_str = str(min_str)
    if min_str.startswith("PT"): # ISO-8601 duration format, e.g. "PT34M12.00S"
        mm = min_str.split("M")[0].replace("PT", "")
        ss = min_str.split("M")[1].replace("S", "") if "M" in min_str else "0"
        return int(float(mm)) + float(ss) / 60
    mm, ss = min_str.split(":") # plain "MM:SS" format
    return int(mm) + int(ss) / 60

print("Fetching Knicks box scores (including DNPs) for all 19 playoff games...")
boxscore_rows = [] # will hold one Knicks-only dataframe per game
for _, game in games_info.iterrows(): # loop through each of the 19 games
    box = boxscoretraditionalv3.BoxScoreTraditionalV3(game_id=game["GAME_ID"], timeout=60).get_data_frames()[0] # full box score
    knicks_box = box[box["teamId"] == TEAM_ID].copy() # keep only the Knicks' own players
    knicks_box["GAME_ID"] = game["GAME_ID"] # make sure the game id is present for the merge below
    boxscore_rows.append(knicks_box)
    time.sleep(0.6) # small pause to avoid overloading the API

boxscore = pd.concat(boxscore_rows, ignore_index=True) # one row per Knicks player per game (played or not)
boxscore = boxscore.rename(columns={"personId": "PLAYER_ID"}) # V3 uses camelCase; standardize the id column name
boxscore["PLAYER_NAME"] = boxscore["firstName"] + " " + boxscore["familyName"] # V3 splits the name into two fields
boxscore["MINUTES_DECIMAL"] = boxscore["minutes"].apply(minutes_to_decimal) # convert to decimal minutes
# comment is the NBA's own official DNP reason (e.g. "DNP - COACH'S DECISION"); blank/NaN means the player actually played
boxscore["STATUS"] = boxscore["comment"].apply(lambda c: c.strip() if isinstance(c, str) and c.strip() else "Yes")

# --- Part C: build the full grid (every player who appeared at least once x all 19 games) ---
all_players = boxscore[["PLAYER_ID", "PLAYER_NAME"]].drop_duplicates() # every player who showed up in any Knicks box score
full_grid = all_players.merge(games_info, how="cross") # cross join -> guarantees exactly 19 rows per player

merged = full_grid.merge( # attach the actual box score data where it exists
    boxscore[["PLAYER_ID", "GAME_ID", "MINUTES_DECIMAL", "points", "reboundsTotal", "assists",
              "turnovers", "steals", "blocks", "foulsPersonal", "STATUS"]],
    on=["PLAYER_ID", "GAME_ID"], how="left"
)

# a player missing entirely from that game's box score (not even listed as inactive/DNP) wasn't with the team that day
missing_row = merged["STATUS"].isna()
merged.loc[missing_row, "STATUS"] = "NWT" # Not With Team
stat_cols = ["MINUTES_DECIMAL", "points", "reboundsTotal", "assists", "turnovers", "steals", "blocks", "foulsPersonal"]
for stat_col in stat_cols:
    merged[stat_col] = merged[stat_col].fillna(0) # DNP/NWT rows get zeroed-out stats

# --- Part D: add the constant columns and put everything in the requested order ---
merged["SEASON"] = SEASON
merged["SEASON_TYPE"] = SEASON_TYPE
merged["TEAM"] = TEAM

players_long = merged.rename(columns={
    "MINUTES_DECIMAL": "MINUTES_PLAYED", "points": "POINTS", "reboundsTotal": "REBOUNDS",
    "assists": "ASSISTS", "turnovers": "TURNOVERS", "steals": "STEALS", "blocks": "BLOCKS", "foulsPersonal": "PERSONAL_FOULS",
})[[
    "SEASON", "SEASON_TYPE", "ROUND_NAME", "TEAM", "OPPONENT", "PLAYER_ID", "PLAYER_NAME", "GAME_NUMBER",
    "MINUTES_PLAYED", "POINTS", "REBOUNDS", "ASSISTS", "TURNOVERS", "STEALS", "BLOCKS", "PERSONAL_FOULS", "STATUS",
]]

# order by player and then by game_number (1..19, chronological across the whole playoff run)
players_long = players_long.sort_values(["PLAYER_NAME", "GAME_NUMBER"]).reset_index(drop=True)

print(f"{len(players_long)} player-game rows found ({players_long['PLAYER_NAME'].nunique()} players x 19 games)")
print(players_long) # preview the dataframe in the terminal

players_long.columns = players_long.columns.str.lower() # convert column headers to lowercase

players_long.to_csv("nyk_players_25_26_playoffs_long_data.csv", index=False) # save as CSV, no row index column

excel_path = "nyk_players_25_26_playoffs_long_data.xlsx" # define the Excel file path
players_long.to_excel(excel_path, index=False, sheet_name="Playoffs") # save as Excel, no row index column

# --- Custom Excel formatting starts here ---
wb = load_workbook(excel_path) # reopen the saved Excel file
ws = wb["Playoffs"] # select the "Playoffs" sheet

plain_font = Font(bold=False) # define a non-bold font style (for data rows)
bold_font = Font(bold=True) # define a bold font style (for header row)
center_align = Alignment(horizontal="center", vertical="center") # define centered alignment (horizontal + vertical)

for row_index, row in enumerate(ws.iter_rows(), start=1): # loop through every row, tracking its row number (starts at 1)
    for cell in row: # loop through every cell in that row
        cell.font = bold_font if row_index == 1 else plain_font # bold only for row 1 (header), plain for the rest
        cell.alignment = center_align # apply centered alignment
        cell.border = None # explicitly remove any border

ws.freeze_panes = "A2" # freeze the top row (everything above row 2 stays fixed when scrolling)

for column_cells in ws.columns: # loop through each column
    max_length = max(len(str(cell.value)) for cell in column_cells) # find the longest value in that column
    col_letter = column_cells[0].column_letter # get the column's letter (A, B, C...)
    ws.column_dimensions[col_letter].width = max_length + 2 # set width = longest text + small padding

wb.save(excel_path) # save the formatted Excel file, overwriting the previous version
print("Saved both CSV and formatted Excel files") # confirm completion